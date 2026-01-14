import os
import openpyxl
import struct

# 类型大小映射
type_size_map = {
    'bool': 1,
    'char': 1,
    'int8_t': 1,
    'uint8_t': 1,
    'int16_t': 2,
    'uint16_t': 2,
    'int32_t': 4,
    'uint32_t': 4,
    'int64_t': 8,
    'uint64_t': 8,
    'float': 4,
    'double': 8
}

def excel_parse(excel_file):
    """解析Excel文件，获取变量信息"""
    variable_detail = {}
    
    # 打开Excel文件
    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    
    # 获取所有表名
    variable_detail['Name'] = workbook.sheetnames
    variable_detail['Number'] = len(variable_detail['Name'])
    
    # 获取每个表的信息
    for i, sheet_name in enumerate(variable_detail['Name']):
        sheet = workbook[sheet_name]
        struct_detail = []
        
        # 从第1行开始读取数据（Excel中没有表头）
        for row in sheet.iter_rows(min_row=1, values_only=True):
            if row[0] is None:  # 如果第一列为空，跳过该行
                continue
            struct_detail.append(row)
        
        variable_detail[f'Var{i+1}'] = struct_detail
    
    workbook.close()
    return variable_detail

def code_generate_type(header_file, variable_detail):
    """生成类型定义"""
    for i in range(variable_detail['Number']):
        sheet_name = variable_detail['Name'][i]
        struct_detail = variable_detail[f'Var{i+1}']
        
        # 写入结构体开始
        header_file.write(f'typedef struct \n')
        header_file.write(f'{{ \n')
        
        # 添加时间戳
        header_file.write(f'  uint64_t timestamp; \n')
        
        # 逐个元素产生代码
        for row in struct_detail:
            type_var, dimension_var, name_var = row
            
            # 声明结构体的一个元素
            header_file.write(f'  {type_var} {name_var}')
            if dimension_var > 1:
                header_file.write(f'[{int(dimension_var)}]')
            header_file.write(f';\n')
        
        # 写入结构体结尾
        header_file.write(f'}}ULog_{sheet_name}_T; \n')
        header_file.write(f'\n')

def code_generate_encoder(topic_name, source_file, variable_detail):
    """生成编码函数"""
    max_variable_size = 0
    source_file.write('__weak static unsigned char default_buffer[];\n\n')
    
    # 声明函数
    function_names = []
    for i in range(variable_detail['Number']):
        sheet_name = variable_detail['Name'][i]
        function_name = f'ULog_Encoder_{sheet_name}'
        function_names.append(function_name)
        source_file.write(f'static unsigned int {function_name}(void * pValue);\n')
    
    # 声明入口初始化函数
    source_file.write(f'static void ULog_{topic_name}_Enter_Init();\n')
    source_file.write('\n')
    
    # 创建变量写入口
    source_file.write('static ULOG_VARIABLE_T ulog_enter_array[] = \n')
    source_file.write('{\n')
    for i in range(variable_detail['Number']):
        sheet_name = variable_detail['Name'][i]
        function_name = function_names[i]
        source_file.write(f'  {"{sheet_name}", {function_name}},\n')
    source_file.write('};\n')
    source_file.write('\n')
    
    # 声明变量
    source_file.write(f'ULOG_VARIABLE_ENTER_T ulog_enter_{topic_name} = \n')
    source_file.write('{\n')
    source_file.write(f'  ULog_{topic_name}_Enter_Init, 0, 0, default_buffer, {variable_detail["Number"]}, ulog_enter_array\n')
    source_file.write('};\n')
    source_file.write('\n')
    
    # 为每个变量生成函数体
    for i in range(variable_detail['Number']):
        sheet_name = variable_detail['Name'][i]
        function_name = function_names[i]
        struct_detail = variable_detail[f'Var{i+1}']
        
        source_file.write(f'static unsigned int {function_name}(void * pValue)\n')
        source_file.write('{\n')
        
        # 声明指针变量
        type_name = f'  ULog_{sheet_name}_T'
        source_file.write(f'{type_name} * pVar = ({type_name} *)pValue;\n')
        source_file.write('  unsigned char * pData;\n')
        source_file.write('\n')
        
        # 消息头的偏移
        counter = 5
        
        # 拷贝时间戳
        source_file.write('  pData = (unsigned char *)&(pVar->timestamp);\n')
        source_file.write(f'  memcpy(&default_buffer[{counter}], pData, 8);\n')
        source_file.write('\n')
        counter += 8
        
        # 逐个元素产生代码
        for row in struct_detail:
            type_var, dimension_var, name_var = row
            
            # 将当前指针指向对应元素
            if dimension_var > 1:
                source_file.write(f'  pData = (unsigned char *)(pVar->{name_var});\n')
            else:
                source_file.write(f'  pData = (unsigned char *)&(pVar->{name_var});\n')
            
            # 确定拷贝长度
            copy_size = int(dimension_var) * type_size_map[type_var]
            
            # 拷贝数据
            source_file.write(f'  memcpy(&default_buffer[{counter}], pData, {copy_size});\n')
            source_file.write('\n')
            
            # 更新缓存索引
            counter += copy_size
        
        # 完善消息头
        # msgSize
        msg_size = counter - 3
        msg_size_bytes = struct.pack('<H', msg_size)
        source_file.write(f'  default_buffer[0] = {msg_size_bytes[0]};\n')
        source_file.write(f'  default_buffer[1] = {msg_size_bytes[1]};\n')
        
        # msgType 'D'
        source_file.write('  default_buffer[2] = 0x44;\n')
        
        # msgId
        msg_id = i
        msg_id_bytes = struct.pack('<H', msg_id)
        source_file.write(f'  default_buffer[3] = {msg_id_bytes[0]};\n')
        source_file.write(f'  default_buffer[4] = {msg_id_bytes[1]};\n')
        
        # 确定所有变量中，所需要的最大缓存
        if max_variable_size < counter:
            max_variable_size = counter
        
        # 输出返回值
        source_file.write('\n')
        source_file.write(f'  return {counter};\n')
        
        # 结束函数
        source_file.write('}\n')
        source_file.write('\n')
    
    # 声明缓存
    source_file.write(f'static unsigned char default_buffer[{max_variable_size}] = {{0}};\n')
    source_file.write('\n')

def code_generate_format_subscribe(topic_name, source_file, variable_detail):
    """生成格式和订阅代码"""
    # 格式及订阅的总字节数
    counter = 0
    
    # 入口初始化函数开始
    source_file.write(f'static void ULog_{topic_name}_Enter_Init()\n')
    source_file.write('{\n')
    
    # 声明格式变量
    source_file.write('  ULOG_MESSAGE_FORMAT_T format[] = \n')
    source_file.write('  {\n')
    
    # 逐个变量的初始化
    for i in range(variable_detail['Number']):
        sheet_name = variable_detail['Name'][i]
        struct_detail = variable_detail[f'Var{i+1}']
        
        # 变量名及时间戳域
        format_string = f'{sheet_name}:uint64_t timestamp;'
        
        # 逐个元素初始化
        for row in struct_detail:
            type_var, dimension_var, name_var = row
            if dimension_var > 1:
                format_string += f'{type_var}[{int(dimension_var)}] {name_var};'
            else:
                format_string += f'{type_var} {name_var};'
        
        # 计算格式大小
        format_size = len(format_string) + 1  # +1 for null terminator
        counter += format_size + 3
        
        # 初始化变量
        source_file.write(f'    {{0x{format_size:04x}, 0x46, "{format_string}"}},
')
    
    # 格式变量结束
    source_file.write('  };\n')
    source_file.write('\n')
    
    # 声明订阅变量
    source_file.write('  ULOG_MESSAGE_SUBSCRIBE_T subscribe[] = \n')
    source_file.write('  {\n')
    
    # 逐个变量的初始化
    for i in range(variable_detail['Number']):
        sheet_name = variable_detail['Name'][i]
        name_length = len(sheet_name) + 1  # +1 for null terminator
        # 修复：移除msg_size的计算，直接使用name_length + 3 + 3
        counter += name_length + 3 + 3  # name_length + 3 bytes header + 3 bytes padding
        
        # 保留msg_size的定义，因为它在生成代码时仍然需要
        msg_size = name_length + 3
        source_file.write(f'    {{0x{msg_size:04x}, 0x41, 0, {i}, "{sheet_name}"}},
')
    
    # 订阅变量结束
    source_file.write('  };\n')
    source_file.write('\n')
    
    # 声明日志头部缓存
    source_file.write(f'  static unsigned char ulog_header_buffer[{counter}] = {{0}};\n')
    source_file.write('  int counter = 0;\n')
    source_file.write('\n')
    
    # 将格式变量拷贝至缓存
    source_file.write(f'  for(int i = 0; i < {variable_detail["Number"]}; i++)\n')
    source_file.write('  {\n')
    source_file.write('    memcpy(&ulog_header_buffer[counter], &format[i], 3);\n')
    source_file.write('    counter = counter + 3;\n')
    source_file.write('    memcpy(&ulog_header_buffer[counter], format[i].format, format[i].msgSize);\n')
    source_file.write('    counter = counter + format[i].msgSize;\n')
    source_file.write('  }\n')
    source_file.write('\n')
    
    # 将订阅变量拷贝至缓存
    source_file.write(f'  for(int i = 0; i < {variable_detail["Number"]}; i++)\n')
    source_file.write('  {\n')
    source_file.write('    memcpy(&ulog_header_buffer[counter], &subscribe[i], 6);\n')
    source_file.write('    counter = counter + 6;\n')
    source_file.write('    memcpy(&ulog_header_buffer[counter], subscribe[i].msgName, subscribe[i].msgSize-3);\n')
    source_file.write('    counter = counter + subscribe[i].msgSize - 3;\n')
    source_file.write('  }\n')
    source_file.write('\n')
    
    # 更新入口
    source_file.write(f'  ulog_enter_{topic_name}.pHeader = ulog_header_buffer;\n')
    source_file.write(f'  ulog_enter_{topic_name}.headerSize = {counter};\n')
    
    # 入口初始化函数结束
    source_file.write('}\n')

def code_generate(topic_name, variable_detail):
    """主函数，协调所有生成过程"""
    # 创建文件
    header_file_path = f'../Topic/{topic_name}.h'
    source_file_path = f'../Topic/{topic_name}.c'
    
    with open(header_file_path, 'w') as header_file, open(source_file_path, 'w') as source_file:
        # 输入头文件的编译代码
        header_file.write(f'#ifndef _{topic_name}_H_\n')
        header_file.write(f'#define _{topic_name}_H_\n')
        header_file.write('\n')
        
        # 头文件包含
        header_file.write('#include "ulog.h"\n')
        header_file.write('\n')
        
        # 头文件声明外部变量
        header_file.write(f'extern ULOG_VARIABLE_ENTER_T ulog_enter_{topic_name};\n')
        header_file.write('\n')
        
        # 源文件包含
        source_file.write(f'#include "{topic_name}.h"\n\n')
        header_file.write('\n')
        
        # 产生变量定义
        code_generate_type(header_file, variable_detail)
        
        # 产生编码函数
        code_generate_encoder(topic_name, source_file, variable_detail)
        
        # 产生格式和订阅代码
        code_generate_format_subscribe(topic_name, source_file, variable_detail)
        
        # 编辑结束部分
        header_file.write('#endif\n')

def main():
    """主函数"""
    
    topic_name = 'Topic'
    config_file = '../Topic/Topic.xlsx'
    
    # 解析Excel文件
    variable_detail = excel_parse(config_file)
    
    # 生成代码
    code_generate(topic_name, variable_detail)
    
    print("代码生成完成！")

if __name__ == '__main__':
    main()